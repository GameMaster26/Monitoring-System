from django.shortcuts import render,redirect
from django.http import HttpResponse
import io
from .models import Patient, History, Treatment,Barangay,Municipality
from django.db.models import Count, F,Sum
import json
from django.db.models.functions import ExtractWeek, ExtractYear, ExtractDay, ExtractMonth
from datetime import datetime, timedelta,date
from dateutil.relativedelta import relativedelta 
from django.utils import timezone
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required,user_passes_test
from django.contrib.auth import get_user_model
from django.contrib.admin.models import LogEntry
from django.template.loader import get_template
#from weasyprint import HTML
import pandas as pd
from openpyxl import Workbook
import openpyxl  # Import the openpyxl library
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font,Alignment
from openpyxl.utils import get_column_letter

from xhtml2pdf import pisa

from io import BytesIO  
import os  
from django.db.models import Q
from django.db.models.functions import TruncDay,TruncDate,TruncMonth, TruncQuarter,TruncYear
from collections import Counter
from urllib.parse import urlencode
from django.http import JsonResponse
from django.template.loader import render_to_string
from calendar import month_name
from django.contrib.auth.models import Group,User
from django.shortcuts import redirect
from django.contrib.auth import logout as auth_logout
from django.contrib import messages
from django.db import models
from django.contrib.gis.geos import GEOSGeometry
from django.core.serializers import serialize


def index(request):
    # Retrieve all totals
    total_patients = Patient.objects.count()
    total_history = History.objects.count()
    total_treatments = Treatment.objects.count()

    # Count male and female patients
    male_patients = Patient.objects.filter(sex='male').count()
    female_patients = Patient.objects.filter(sex='female').count()

    # Calculate counts for different animal bites
    dog_bites = History.objects.filter(source_of_exposure='Dog').count()
    cat_bites = History.objects.filter(source_of_exposure='Cat').count()
    
    # Perform a query to count the occurrences of each municipality of exposure
    municipality_exposures = History.objects.values('muni_id__muni_name').annotate(count=Count('muni_id')).order_by('-count').first()

    # Determine which animal has the highest count
    animal_counts = History.objects.values('source_of_exposure').annotate(count=Count('source_of_exposure')).order_by('-count').first()
    most_cases_animal = animal_counts

    gender_counts = Patient.objects.values('sex').annotate(count=Count('sex')).order_by('-count').first()

    histories = History.objects.all()  # Retrieve all Histories from the database

    # Count the number of male and female patients
    gender = Patient.objects.values('sex').annotate(count=Count('sex')).order_by('sex')
    
    gen = [data['sex'].capitalize() for data in gender]
    dataGender = [data['count'] for data in gender]

    # Count the number animal source of exposure
    source_exposure = History.objects.values('source_of_exposure').annotate(count=Count('source_of_exposure')).order_by('source_of_exposure')
    
    animal = [data['source_of_exposure'].capitalize() for data in source_exposure]
    dataAnimal = [data['count'] for data in source_exposure]

    # Count occurrences of each municipality using Django ORM
    municipality_counts = History.objects.values('muni_id__muni_name').annotate(count=Count('muni_id')).order_by('muni_id__muni_name')

    # Prepare data for chart
    municipalities = [data['muni_id__muni_name'] for data in municipality_counts]
    municipality_case_counts = [data['count'] for data in municipality_counts]

    # Group cases by month
    monthly_cases = History.objects.annotate(month=TruncMonth('date_registered')).values('month').annotate(count=Count('history_id')).order_by('month')

    months = [data['month'].strftime('%B') for data in monthly_cases]
    case_counts = [data['count'] for data in monthly_cases]

    # Calculate the number of cases per week
    weekly_cases = History.objects.annotate(
        week=ExtractWeek('date_registered'),
        year=ExtractYear('date_registered')
    ).values('year', 'week').annotate(count=Count('history_id')).order_by('year', 'week')

    # Prepare the data for the chart with start dates of weeks
    weeks = []
    weekly_case_counts = []
    for entry in weekly_cases:
        year = entry['year']
        week = entry['week']
        start_date = datetime.strptime(f'{year}-W{week}-1', "%Y-W%U-%w")  # Ensure the week starts from Sunday
        week_label = start_date.strftime('%Y-%m-%d')
        weeks.append(week_label)
        weekly_case_counts.append(entry['count'])

    # Calculate the number of cases per day
    daily_cases = (
        History.objects
        .annotate(day=TruncDate('date_registered'))
        .values('day')
        .annotate(count=Count('history_id'))
        .order_by('day')
    )

    days = []
    daily_case_counts = []

    for data in daily_cases:
        day = data['day']
        count = data['count']
        
        if day is not None:
            days.append(day.strftime('%Y-%m-%d'))
        else:
            days.append('Unknown')  # Handle None case with a fallback
        
        daily_case_counts.append(count)

    # Calculate the number of cases per quarter
    quarterly_cases = (
        History.objects
        .annotate(quarter=TruncQuarter('date_registered'))
        .values('quarter')
        .annotate(count=Count('history_id'))
        .order_by('quarter')
    )

    # Prepare quarters and quarterly counts
    quarters = []
    quarterly_case_counts = []
    
    for data in quarterly_cases:
        quarter_str = f"Quarter {((data['quarter'].month - 1) // 3) + 1} {data['quarter'].year}"
        quarters.append(quarter_str)
        quarterly_case_counts.append(data['count'])

        
    # Calculate the number of cases per year
    annual_cases = (
        History.objects
        .annotate(year=TruncYear('date_registered'))
        .values('year')
        .annotate(count=Count('history_id'))
        .order_by('year')
    )

    years = [data['year'].strftime('%Y') for data in annual_cases]
    annual_case_counts = [data['count'] for data in annual_cases]


    total_cases = History.objects.count()

    # Get the start and end dates from the GET parameters
    start_date = request.GET.get('startDate')
    end_date = request.GET.get('endDate')

    # Filter history objects based on date range
    if start_date and end_date:
        # Convert string dates to datetime objects
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')

        # Filter heatmap data based on the selected date range
        heatmap_data = History.objects.filter(
            date_registered__range=(start_date, end_date)  # Ensure to use the correct field
        ).values('latitude', 'longitude').annotate(count=Count('history_id'))

        # Filter human rabies cases based on the selected date range
        rabies_heatmap_data = History.objects.filter(
            date_registered__range=(start_date, end_date),
            human_rabies=True
        ).values('latitude', 'longitude').annotate(count=Count('history_id'))

        # Get the total rabies cases within the date range
        total_rabies_cases = History.objects.filter(
            date_registered__range=(start_date, end_date),
            human_rabies=True
        ).count()

        # Get the total cases within the date range
        total_cases = History.objects.filter(
            date_registered__range=(start_date, end_date)
        ).count()

    else:
        # If no dates provided, get all history
        heatmap_data = History.objects.values('latitude', 'longitude').annotate(count=Count('history_id'))
        total_cases = History.objects.count()  # Total cases without filter

        # For human rabies data without date filter
        rabies_heatmap_data = History.objects.filter(human_rabies=True).values('latitude', 'longitude').annotate(count=Count('history_id'))
        total_rabies_cases = History.objects.filter(human_rabies=True).count()

    # Prepare the heatmap data for JSON
    heatmap_data = [[entry['latitude'], entry['longitude'], entry['count']] for entry in heatmap_data]

    # Prepare the rabies heatmap data for JSON
    rabies_heatmap_data = [[entry['latitude'], entry['longitude'], entry['count']] for entry in rabies_heatmap_data]

    # Get human rabies cases by municipality
    rabies_municipality_counts = History.objects.filter(human_rabies=True).values('muni_id__muni_name').annotate(count=Count('muni_id')).order_by('-count')

    # Prepare data for human rabies by municipality
    rabies_municipalities = [data['muni_id__muni_name'] for data in rabies_municipality_counts]
    rabies_case_counts = [data['count'] for data in rabies_municipality_counts]

    context = {
        'total_patients': total_patients,
        'total_history': total_history,
        'total_treatments': total_treatments,
        'male_patients': male_patients,
        'female_patients': female_patients,
        'dog_bites': dog_bites,
        'cat_bites': cat_bites,
        'most_cases_animal': most_cases_animal,
        'municipality': municipality_exposures,
        'gender_counts': gender_counts,
        'histories': histories,
        'gen': gen,
        'dataGender': dataGender,
        'animal': animal,
        'dataAnimal': dataAnimal,
        'municipalities': municipalities,
        'municipality_case_counts': municipality_case_counts,
        'months': months,
        'case_counts': case_counts,
        'weeks': weeks,
        'weekly_case_counts': weekly_case_counts,
        'days': days,
        'daily_case_counts': daily_case_counts,
        'quarters': quarters,
        'quarterly_case_counts': quarterly_case_counts,
        'years': years,
        'annual_case_counts': annual_case_counts,
        'total_cases': total_cases,
        'heatmap_data': json.dumps(heatmap_data),
        'total_rabies_cases': total_rabies_cases,
        'rabies_municipalities': rabies_municipalities,
        'rabies_case_counts': rabies_case_counts,
        'rabies_heatmap_data': json.dumps(rabies_heatmap_data),
    }

    return render(request, 'monitoring/index.html', context)

def choropleth_map(request):
    
    #Choropleth Map
    barangay_density = History.objects.values('brgy_id').annotate(patient_count=Count('patient_id')).order_by('brgy_id')

    # Convert this to a dictionary where the key is brgy_id and the value is the patient_count
    density_dict = {str(d['brgy_id']): d['patient_count'] for d in barangay_density}

    # Get human rabies cases by municipality
    rabies_municipality_counts = History.objects.filter(human_rabies=True).values('muni_id__muni_name').annotate(count=Count('muni_id')).order_by('-count')
    
    # Total rabies cases
    total_rabies_cases = History.objects.filter(human_rabies=True).count()

    # Prepare data for human rabies by municipality
    rabies_municipalities = [data['muni_id__muni_name'] for data in rabies_municipality_counts]
    rabies_case_counts = [data['count'] for data in rabies_municipality_counts]

    # Prepare heatmap data for human rabies cases
    rabies_heatmap_data = History.objects.filter(human_rabies=True).values('latitude', 'longitude').annotate(count=Count('history_id'))

    # Prepare the heatmap data for JSON
    rabies_heatmap_data = [[entry['latitude'], entry['longitude'], entry['count']] for entry in rabies_heatmap_data]
    
    # Fetch all municipalities
    all_municipalities = Municipality.objects.all()

    # Get the selected municipality and barangay search from the POST parameters
    if request.method == 'POST':
        selected_municipality = request.POST.get('municipality', None)
        barangay_search = request.POST.get('barangay', '').strip()  # Get barangay search input and remove extra spaces
    else:
        selected_municipality = "ALL"
        barangay_search = ""  # Default to an empty search

    # Initialize variables for summary
    total_barangays = 0
    total_patients = 0
    barangay_summary = []
    municipality_summary = []

    if selected_municipality == "ALL" or selected_municipality is None:
        # If "ALL" is selected, show the municipality-level summary
        municipality_summary = (
            History.objects.values('muni_id__muni_name')  # Group by municipality
            .annotate(total_barangays=Count('brgy_id', distinct=True),  # Count distinct barangays
                      total_cases=Count('history_id'))  # Count total cases (patients)
            .order_by('muni_id__muni_name')
        )

        # Calculate the total number of barangays and patients across all municipalities
        total_barangays = sum(record['total_barangays'] for record in municipality_summary)
        total_patients = sum(record['total_cases'] for record in municipality_summary)

    else:
        # Get all barangays for the selected municipality
        barangays_in_municipality = Barangay.objects.filter(muni_id__muni_name=selected_municipality)
        
        # Filter barangays based on the search input (partial match)
        if barangay_search:
            barangays_in_municipality = barangays_in_municipality.filter(brgy_name__icontains=barangay_search)

        # Get barangay-level case summary
        cases_in_barangays = History.objects.filter(muni_id__muni_name=selected_municipality) \
            .values('brgy_id__brgy_name') \
            .annotate(total_patients=Count('history_id')) \
            .order_by('brgy_id__brgy_name')

        # Create a dictionary for easy lookup
        cases_dict = {record['brgy_id__brgy_name']: record['total_patients'] for record in cases_in_barangays}

        # Populate barangay summary: if no cases, set total patients to 0
        barangay_summary = [
            {
                'brgy_name': barangay.brgy_name,
                'total_patients': cases_dict.get(barangay.brgy_name, 0)  # Get patients, default to 0 if not found
            }
            for barangay in barangays_in_municipality
        ]

        # Calculate totals
        total_barangays = barangays_in_municipality.count()
        total_patients = sum(record['total_patients'] for record in barangay_summary)

        # Summary for the selected municipality
        municipality_summary = [
            {
                'muni_id__muni_name': selected_municipality,
                'total_barangays': total_barangays,
                'total_cases': total_patients,
            }
        ]

    # Pass the data to the template as part of the context
    context = {
        'total_rabies_cases': total_rabies_cases,
        'rabies_case_counts': rabies_case_counts,
        'density_dict': density_dict,
        'rabies_heatmap_data': json.dumps(rabies_heatmap_data),
        'barangay_summary': barangay_summary,
        'municipality_summary': municipality_summary,
        'municipalities': all_municipalities,
        'selected_municipality': selected_municipality,
        'barangay_search': barangay_search,  # Pass the search input to the template
        'total_barangays': total_barangays,
        'total_patients': total_patients,
        'all_municipalities': all_municipalities,
    }

    return render(request, 'monitoring/choropleth.html', context)



def choro(request):
    # Get barangay densities: Count the number of patients per barangay
    barangay_density = History.objects.values('brgy_id').annotate(patient_count=Count('patient_id')).order_by('brgy_id')

    # Convert this to a dictionary where the key is brgy_id and the value is the patient_count
    density_dict = {str(d['brgy_id']): d['patient_count'] for d in barangay_density}

    # Get human rabies cases by municipality
    rabies_municipality_counts = History.objects.filter(human_rabies=True).values('muni_id__muni_name').annotate(count=Count('muni_id')).order_by('-count')
    
    # Total rabies cases
    total_rabies_cases = History.objects.filter(human_rabies=True).count()

    # Prepare data for human rabies by municipality
    rabies_municipalities = [data['muni_id__muni_name'] for data in rabies_municipality_counts]
    rabies_case_counts = [data['count'] for data in rabies_municipality_counts]

    # Prepare heatmap data for human rabies cases
    rabies_heatmap_data = History.objects.filter(human_rabies=True).values('latitude', 'longitude').annotate(count=Count('history_id'))

    # Prepare the heatmap data for JSON
    rabies_heatmap_data = [[entry['latitude'], entry['longitude'], entry['count']] for entry in rabies_heatmap_data]

    # Pass the data to the template as part of the context
    context = {
        'density_dict': density_dict,  # For barangay patient density
        'rabies_case_counts': rabies_case_counts,
        'total_rabies_cases': total_rabies_cases,
        'rabies_heatmap_data': json.dumps(rabies_heatmap_data),

    }
    return render(request, 'monitoring/choro.html', context)



def admin_redirect(request):
    return redirect('/admin/')

def logout(request):
    auth_logout(request) 
    return redirect('/admin/login/')



# Function to calculate the age of the patient
def calculate_age(birthday):
    today = date.today()
    return today.year - birthday.year - ((today.month, today.day) < (birthday.month, birthday.day))

def staff_or_superuser_required(view_func):
    def _wrapped_view_func(request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.warning(request, "Only staff members or superusers can access this page.")
            return redirect('/admin/')
        return view_func(request, *args, **kwargs)
    return login_required(_wrapped_view_func)


def superuser_required(view_func):
    def _wrapped_view_func(request, *args, **kwargs):
        if not request.user.is_superuser:
            messages.warning(request, "Only superusers or admins can access this page.")
            return redirect('/admin/')  # Replace 'some_view_name' with the name of the view you want to redirect to
        return view_func(request, *args, **kwargs)
    return login_required(_wrapped_view_func)

#@superuser_required
#@staff_or_superuser_required
@staff_member_required
@login_required
def overview(request):
    # Get the start and end dates from the GET parameters
    start_date = request.GET.get('startDate')
    end_date = request.GET.get('endDate')

    # Filter history objects based on date range
    if start_date and end_date:
        # Convert string dates to datetime objects
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')

        # Filter heatmap data based on the selected date range
        heatmap_data = History.objects.filter(
            date_registered__range=(start_date, end_date)  # Ensure to use the correct field
        ).values('latitude', 'longitude').annotate(count=Count('history_id'))

        # Get the total cases within the date range
        total_cases = heatmap_data.count()

        # Filter human rabies cases based on the selected date range
        total_rabies_cases = History.objects.filter(
            human_rabies=True,
            date_registered__range=(start_date, end_date)
        ).count()

        # Filter heatmap data for human rabies cases based on date range
        rabies_heatmap_data = History.objects.filter(
            human_rabies=True,
            date_registered__range=(start_date, end_date)
        ).values('latitude', 'longitude').annotate(count=Count('history_id'))

        # Get human rabies cases by municipality for the date range
        rabies_municipality_counts = History.objects.filter(
            human_rabies=True,
            date_registered__range=(start_date, end_date)
        ).values('muni_id__muni_name').annotate(count=Count('muni_id')).order_by('-count')

    else:
        # If no dates provided, get all history
        heatmap_data = History.objects.values('latitude', 'longitude').annotate(count=Count('history_id'))
        total_cases = History.objects.count()  # Total cases without filter

        # Count total human rabies cases
        total_rabies_cases = History.objects.filter(human_rabies=True).count()

        # Prepare heatmap data for human rabies cases
        rabies_heatmap_data = History.objects.filter(human_rabies=True).values('latitude', 'longitude').annotate(count=Count('history_id'))

        # Get human rabies cases by municipality
        rabies_municipality_counts = History.objects.filter(human_rabies=True).values('muni_id__muni_name').annotate(count=Count('muni_id')).order_by('-count')

    # Prepare the heatmap data for JSON
    heatmap_data = [[entry['latitude'], entry['longitude'], entry['count']] for entry in heatmap_data]

    # Prepare the rabies heatmap data for JSON
    rabies_heatmap_data = [[entry['latitude'], entry['longitude'], entry['count']] for entry in rabies_heatmap_data]

    # Prepare data for human rabies by municipality
    rabies_municipalities = [data['muni_id__muni_name'] for data in rabies_municipality_counts]
    rabies_case_counts = [data['count'] for data in rabies_municipality_counts]

    # Count occurrences of each municipality using Django ORM
    municipality_counts = History.objects.values('muni_id__muni_name').annotate(count=Count('muni_id')).order_by('muni_id__muni_name')

    # Prepare data for chart
    municipalities = [data['muni_id__muni_name'] for data in municipality_counts]
    municipality_case_counts = [data['count'] for data in municipality_counts]

    # Count the number of male and female patients
    gender = Patient.objects.values('sex').annotate(count=Count('sex')).order_by('sex')
    
    gen = [data['sex'].capitalize() for data in gender]
    dataGender = [data['count'] for data in gender]

    # Count the number of animal sources of exposure
    source_exposure = History.objects.values('source_of_exposure').annotate(count=Count('source_of_exposure')).order_by('source_of_exposure')
    
    animal = [data['source_of_exposure'].capitalize() for data in source_exposure]
    dataAnimal = [data['count'] for data in source_exposure]

    # Group cases by month
    monthly_cases = History.objects.annotate(month=TruncMonth('date_registered')).values('month').annotate(count=Count('history_id')).order_by('month')

    months = [data['month'].strftime('%B') for data in monthly_cases]
    case_counts = [data['count'] for data in monthly_cases]

    # Calculate the number of cases per week
    weekly_cases = History.objects.annotate(
        week=ExtractWeek('date_registered'),
        year=ExtractYear('date_registered')
    ).values('year', 'week').annotate(count=Count('history_id')).order_by('year', 'week')

    # Prepare the data for the chart with start dates of weeks
    weeks = []
    weekly_case_counts = []
    for entry in weekly_cases:
        year = entry['year']
        week = entry['week']
        start_date = datetime.strptime(f'{year}-W{week}-1', "%Y-W%U-%w")  # Ensure the week starts from Sunday
        week_label = start_date.strftime('%Y-%m-%d')
        weeks.append(week_label)
        weekly_case_counts.append(entry['count'])

    # Calculate the number of cases per day
    daily_cases = (
        History.objects
        .annotate(day=TruncDate('date_registered'))
        .values('day')
        .annotate(count=Count('history_id'))
        .order_by('day')
    )

    days = []
    daily_case_counts = []

    for data in daily_cases:
        day = data['day']
        count = data['count']
        
        if day is not None:
            days.append(day.strftime('%Y-%m-%d'))
        else:
            days.append('Unknown')  # Handle None case with a fallback
        
        daily_case_counts.append(count)

    # Calculate the number of cases per quarter
    quarterly_cases = (
        History.objects
        .annotate(quarter=TruncQuarter('date_registered'))
        .values('quarter')
        .annotate(count=Count('history_id'))
        .order_by('quarter')
    )

    # Prepare quarters and quarterly counts
    quarters = []
    quarterly_case_counts = []
    
    for data in quarterly_cases:
        quarter_str = f"Quarter {((data['quarter'].month - 1) // 3) + 1} {data['quarter'].year}"
        quarters.append(quarter_str)
        quarterly_case_counts.append(data['count'])

    # Calculate the number of cases per year
    annual_cases = (
        History.objects
        .annotate(year=TruncYear('date_registered'))
        .values('year')
        .annotate(count=Count('history_id'))
        .order_by('year')
    )

    years = [data['year'].strftime('%Y') for data in annual_cases]
    annual_case_counts = [data['count'] for data in annual_cases]

    # Pass all the data to the template
    context = {
        'heatmap_data': json.dumps(heatmap_data),       
        'total_cases': total_cases,                        
        'total_rabies_cases': total_rabies_cases,          
        'rabies_municipalities': rabies_municipalities,    
        'rabies_case_counts': rabies_case_counts,          
        'rabies_heatmap_data': json.dumps(rabies_heatmap_data), 
        'municipalities': municipalities,
        'municipality_case_counts': municipality_case_counts,
        'gen': gen,
        'dataGender': dataGender,
        'animal': animal,
        'dataAnimal': dataAnimal,
        'days': days,
        'daily_case_counts': daily_case_counts,
        'weeks': weeks,
        'weekly_case_counts': weekly_case_counts,
        'months': months,
        'case_counts': case_counts,
        'quarters': quarters,
        'quarterly_case_counts': quarterly_case_counts,
        'years': years,
        'annual_case_counts': annual_case_counts,
    }

    return render(request, 'monitoring/overview.html', context)




# Function to calculate age based on birthday
def calculate_age(birthday):
    today = datetime.today()
    age = today.year - birthday.year - ((today.month, today.day) < (birthday.month, birthday.day))
    return age

@login_required
def reports(request):
    user = request.user

    municipality_map = {
        "MAR": "Maripipi",
        "KAW": "Kawayan",
        "NAV": "Naval",
        "CAIB": "Caibiran"
    }

    # Check if the user is a superuser and if the municipality is NAV
    if user.is_superuser and user.code == "NAV":
        municipality_name = "BPH"
    else:
        municipality_name = municipality_map.get(user.code, "Naval")

    # Get the selected quarter from the request (default to 2nd quarter)
    selected_quarter = request.GET.get('quarter', '1')
    year = 2024

    # Define quarter date ranges
    quarter_ranges = {
        '1': (date(year, 1, 1), date(year, 3, 31)),
        '2': (date(year, 4, 1), date(year, 6, 30)),
        '3': (date(year, 7, 1), date(year, 9, 30)),
        '4': (date(year, 10, 1), date(year, 12, 31)),
    }

    start_date, end_date = quarter_ranges[selected_quarter]

    data = []
    total_male = 0
    total_female = 0
    total_all = 0
    total_age_15_below = 0
    total_age_above_15 = 0
    
    # Initialize percentages
    total_sex_percentage = 0

    # Initialize counts for animal bites
    total_animal_bite_I = 0
    total_animal_bite_II = 0
    total_animal_bite_III = 0
    total_category_percentage = 0

    # Add PEP treatment logic and accumulate totals
    total_tcv_given = 0
    total_hrig_given = 0
    total_erig_given = 0  # Initialize ERIG total

    # Percentage variables
    total_tcv_percentage = 0
    total_hrig_percentage = 0
    total_rig_percentage = 0

    # Initialize counts for animal types
    total_dog_bites = 0
    total_cat_bites = 0
    total_other_bites = 0
    total_animal_type_percentage = 0
    total_animal_bite_I_percentage = 0
    total_animal_bite_II_percentage = 0
    total_animal_bite_III_percentage = 0

    # Initialize counters for immunized and unimmunized patients
    total_immunized = 0
    total_unimmunized = 0


    if user.is_superuser:
        abtcs = User.objects.filter(is_superuser=False).distinct()
        for abtc_user in abtcs:
            male_count = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='male'
            ).count()
            female_count = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='female'
            ).count()

            # Filter and calculate age groups based on birthday and registration date within the selected quarter
            patients = Patient.objects.filter(
                user=abtc_user,  # Filter by ABTC user (for superuser view)
                histories__date_registered__range=(start_date, end_date)  # Filter by registration date within the quarter
            ).distinct()
            age_15_below_count = sum(1 for patient in patients if calculate_age(patient.birthday) <= 15)
            age_above_15_count = sum(1 for patient in patients if calculate_age(patient.birthday) > 15)



            # Initialize animal bite counts for each user
            user_animal_bite_I = 0
            user_animal_bite_II = 0
            user_animal_bite_III = 0

            # Count animal bites by category for the specific user
            animal_bite_counts = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date)
            ).values('category_of_exposure').annotate(count=models.Count('category_of_exposure'))

            for count in animal_bite_counts:
                if count['category_of_exposure'] == 'I':
                    user_animal_bite_I = count['count']
                elif count['category_of_exposure'] == 'II':
                    user_animal_bite_II = count['count']
                elif count['category_of_exposure'] == 'III':
                    user_animal_bite_III = count['count']



            # Count PEP treatments for the specific user
            tcv_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                tcv_given__range=(start_date, end_date)
            ).count()

            # Count HRIG treatments
            hrig_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                hrig_given__range=(start_date, end_date)
            ).count()

            # Count ERIG treatments (add logic for ERIG if it's defined in your Treatment model)
            erig_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                rig_given__range=(start_date, end_date)  # Assuming this field exists
            ).count()


            # Count bites for each animal type for the specific user
            animal_type_counts = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date)
            ).values('source_of_exposure').annotate(count=models.Count('source_of_exposure'))

            # Initialize local counts
            user_dog_bites = 0
            user_cat_bites = 0
            user_other_bites = 0

            for count in animal_type_counts:
                if count['source_of_exposure'] == 'Dog':
                    user_dog_bites = count['count']
                elif count['source_of_exposure'] == 'Cat':
                    user_cat_bites = count['count']
                elif count['source_of_exposure'] == 'Others':
                    user_other_bites = count['count']

            # Count immunized and unimmunized patients for this specific user
            user_immunized_count = History.objects.filter(
                patient_id__user=abtc_user,
                immunization_status='Immunized',
                date_registered__range=(start_date, end_date)
            ).count()
            
            user_unimmunized_count = History.objects.filter(
                patient_id__user=abtc_user,
                immunization_status='Unimmunized',
                date_registered__range=(start_date, end_date)
            ).count()

            # Count human rabies cases for the specific user
            user_human_rabies_count = History.objects.filter(
                patient_id__user=abtc_user,
                human_rabies=True,
                date_registered__range=(start_date, end_date)
            ).count()
                    
            # Accumulate totals
            total_immunized += user_immunized_count
            total_unimmunized += user_unimmunized_count
            
            # Accumulate totals for each type
            total_dog_bites += user_dog_bites
            total_cat_bites += user_cat_bites
            total_other_bites += user_other_bites

            # Accumulate totals for PEP treatments
            total_tcv_given += tcv_count
            total_hrig_given += hrig_count
            total_erig_given += erig_count

            # Add to total counts
            total_count = male_count + female_count
            total_male += male_count
            total_female += female_count
            total_all += total_count
            total_age_15_below += age_15_below_count
            total_age_above_15 += age_above_15_count
            # Accumulate totals for reporting
            total_animal_bite_I += user_animal_bite_I
            total_animal_bite_II += user_animal_bite_II
            total_animal_bite_III += user_animal_bite_III

            data.append({
                'barangay': f"{municipality_map.get(abtc_user.code, 'Unknown')}-ABTC",
                'data_male': male_count,
                'data_female': female_count,
                'data_total': total_count,
                'age_15_below': age_15_below_count,
                'age_above_15': age_above_15_count,
                'age_total': age_15_below_count + age_above_15_count,
                'total_animal_bite_I':user_animal_bite_I,
                'total_animal_bite_II':user_animal_bite_II,
                'total_animal_bite_III':user_animal_bite_III,
                'total_animal':user_animal_bite_I + user_animal_bite_II + user_animal_bite_III,
                'total_tcv_given': tcv_count,
                'total_hrig_given': hrig_count,
                'total_erig_given': erig_count,
                'total_dog_bites': user_dog_bites,
                'total_cat_bites': user_cat_bites,
                'total_other_bites': user_other_bites,
                'total_animal_bites': user_dog_bites + user_cat_bites + user_other_bites,
                'immunized_count': user_immunized_count,  # Change made here
                'unimmunized_count': user_unimmunized_count,  # Change made here
                'human_rabies_count':user_human_rabies_count,

            })
    else:
        # Fetch patients for non-superuser
        patients = Patient.objects.filter(user=user)
        barangays = Barangay.objects.filter(patients_brgy__in=patients).distinct()

        for barangay in barangays:
            male_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='male'
            ).count()
            female_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='female'
            ).count()

            # Filter and calculate age groups based on birthday
            patients = Patient.objects.filter(
                brgy_id=barangay,
                user=user,
                histories__date_registered__range=(start_date, end_date)  # Filter by registration date within the quarter
            ).distinct()
            age_15_below_count = sum(1 for patient in patients if calculate_age(patient.birthday) <= 15)
            age_above_15_count = sum(1 for patient in patients if calculate_age(patient.birthday) > 15)

            # Initialize category counts for each barangay
            barangay_animal_bite_I = 0
            barangay_animal_bite_II = 0
            barangay_animal_bite_III = 0

            # Count animal bites by category for the specific barangay
            animal_bite_counts = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).values('category_of_exposure').annotate(count=models.Count('category_of_exposure'))

            tcv_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                tcv_given__range=(start_date, end_date)
            ).count()

            hrig_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                hrig_given__range=(start_date, end_date)
            ).count()

            # ERIG count logic
            erig_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                rig_given__range=(start_date, end_date)
            ).count()

            # Accumulate totals for PEP treatments
            total_tcv_given += tcv_count
            total_hrig_given += hrig_count
            total_erig_given += erig_count

            for count in animal_bite_counts:
                if count['category_of_exposure'] == 'I':
                    barangay_animal_bite_I = count['count']
                elif count['category_of_exposure'] == 'II':
                    barangay_animal_bite_II = count['count']
                elif count['category_of_exposure'] == 'III':
                    barangay_animal_bite_III = count['count']




            # Count bites for each animal type for the specific barangay
            animal_type_counts = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).values('source_of_exposure').annotate(count=models.Count('source_of_exposure'))

            # Initialize local counts
            barangay_dog_bites = 0
            barangay_cat_bites = 0
            barangay_other_bites = 0

            for count in animal_type_counts:
                if count['source_of_exposure'] == 'Dog':
                    barangay_dog_bites = count['count']
                elif count['source_of_exposure'] == 'Cat':
                    barangay_cat_bites = count['count']
                elif count['source_of_exposure'] == 'Others':
                    barangay_other_bites = count['count']

            # Count immunized and unimmunized patients for this specific barangay
            barangay_immunized_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                immunization_status='Immunized',
                date_registered__range=(start_date, end_date)
            ).count()
            
            barangay_unimmunized_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                immunization_status='Unimmunized',
                date_registered__range=(start_date, end_date)
            ).count()

            # Count human rabies cases for the specific barangay
            barangay_human_rabies_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                human_rabies=True,
                date_registered__range=(start_date, end_date)
            ).count()

             # Accumulate totals
            total_immunized += barangay_immunized_count
            total_unimmunized += barangay_unimmunized_count

            # Accumulate totals
            total_dog_bites += barangay_dog_bites
            total_cat_bites += barangay_cat_bites
            total_other_bites += barangay_other_bites

            # Accumulate totals
            total_dog_bites += barangay_dog_bites
            total_cat_bites += barangay_cat_bites
            total_other_bites += barangay_other_bites

            # Accumulate totals for reporting
            total_animal_bite_I += barangay_animal_bite_I
            total_animal_bite_II += barangay_animal_bite_II
            total_animal_bite_III += barangay_animal_bite_III

            # Add to total counts
            total_count = male_count + female_count
            total_male += male_count
            total_female += female_count
            total_all += total_count
            total_age_15_below += age_15_below_count
            total_age_above_15 += age_above_15_count

            data.append({
                'barangay': barangay.brgy_name,
                'data_male': male_count,
                'data_female': female_count,
                'data_total': total_count,
                'age_15_below': age_15_below_count,
                'age_above_15': age_above_15_count,
                'age_total': age_15_below_count + age_above_15_count,
                'total_animal_bite_I': barangay_animal_bite_I,
                'total_animal_bite_II': barangay_animal_bite_II,
                'total_animal_bite_III': barangay_animal_bite_III,
                'total_animal': barangay_animal_bite_I + barangay_animal_bite_II + barangay_animal_bite_III,
                'total_tcv_given': tcv_count,
                'total_hrig_given': hrig_count,
                'total_erig_given': erig_count,
                'total_dog_bites': barangay_dog_bites,
                'total_cat_bites': barangay_cat_bites,
                'total_other_bites': barangay_other_bites,
                'total_animal_bites': barangay_dog_bites + barangay_cat_bites + barangay_other_bites,
                'immunized_count': barangay_immunized_count,  # Change made here
                'unimmunized_count': barangay_unimmunized_count,  # Change made here
                'human_rabies_count':barangay_human_rabies_count,

            })
    
    # After computing total counts
    if total_all > 0:
        male_percentage = (total_male / total_all) * 100
        female_percentage = (total_female / total_all) * 100
        total_sex_percentage = (male_percentage + female_percentage)
        age_15_below_percentage = (total_age_15_below / total_all) * 100
        age_above_15_percentage = (total_age_above_15 / total_all) * 100
        total_age_percentage = (total_age_15_below + total_age_above_15) / total_all * 100
        total_animal_bite_I_percentage = (total_animal_bite_I / total_all ) * 100
        total_animal_bite_II_percentage = (total_animal_bite_II / total_all ) * 100
        total_animal_bite_III_percentage = (total_animal_bite_III / total_all ) * 100
        total_category_percentage = (total_animal_bite_I_percentage + total_animal_bite_II_percentage + total_animal_bite_III_percentage)
        total_tcv_percentage = (total_tcv_given / total_all) * 100
        total_hrig_percentage = (total_hrig_given / total_all) * 100
        total_rig_percentage = (total_erig_given / total_all) * 100  
        dog_bite_percentage = (total_dog_bites / total_all) * 100
        cat_bite_percentage = (total_cat_bites / total_all) * 100
        other_bite_percentage = (total_other_bites / total_all) * 100
        total_animal_type_percentage = dog_bite_percentage + cat_bite_percentage + other_bite_percentage
        # Calculate percentages for immunized and unimmunized
        if total_immunized + total_unimmunized > 0:
            immunized_percentage = (total_immunized / (total_immunized + total_unimmunized)) * 100
            unimmunized_percentage = (total_unimmunized / (total_immunized + total_unimmunized)) * 100
        else:
            immunized_percentage = 0
            unimmunized_percentage = 0
    else:
        male_percentage = female_percentage = age_15_below_percentage = age_above_15_percentage = total_age_percentage = 0
        total_tcv_percentage = total_hrig_percentage = total_rig_percentage = 0
        dog_bite_percentage = cat_bite_percentage = other_bite_percentage = 0
        immunized_percentage = 0
        unimmunized_percentage = 0


    # Calculate overall total for different categories
    overall_total = sum(entry.get('data_total', 0) for entry in data)
    overall_total_tcv = sum(entry.get('total_tcv_given', 0) for entry in data)
    overall_total_hrig = sum(entry.get('total_hrig_given', 0) for entry in data)
    overall_total_erig = sum(entry.get('total_erig_given', 0) for entry in data)

    # Now calculate percentages for each entry
    for entry in data:
        # Total percentage of data_total
        entry['percent_total'] = round((entry['data_total'] / overall_total) * 100, 1) if overall_total > 0 else 0
        
        # Percentages for TCV, HRIG, ERIG, calculated relative to overall totals
        entry['percent_tcv'] = round((entry.get('total_tcv_given', 0) / overall_total_tcv) * 100, 1) if overall_total_tcv > 0 else 0
        entry['percent_hrig'] = round((entry.get('total_hrig_given', 0) / overall_total_hrig) * 100, 1) if overall_total_hrig > 0 else 0
        entry['percent_erig'] = round((entry.get('total_erig_given', 0) / overall_total_erig) * 100, 1) if overall_total_erig > 0 else 0

    
    # Calculate the sum of all percent totals
    total_percent = sum(entry['percent_total'] for entry in data)
    total_tcv_percent = sum(entry['percent_tcv'] for entry in data)
    total_hrig_percent = sum(entry['percent_hrig'] for entry in data)
    total_erig_percent = sum(entry['percent_erig'] for entry in data)
    
    total_human_rabies = sum(entry.get('human_rabies_count', 0) for entry in data)  # Add this line

    


    context = {
        'municipality_name': municipality_name,
        'selected_quarter': selected_quarter,
        'barangay_list': [d['barangay'] for d in data],
        'data': data,
        'total_male': total_male,
        'total_female': total_female,
        'total_all': total_all,
        'total_age_15_below': total_age_15_below,
        'total_age_above_15': total_age_above_15,
        'male_percentage': round(male_percentage, 1),
        'female_percentage': round(female_percentage, 1),
        'total_sex_percentage':round(total_sex_percentage),
        'age_15_below_percentage': round(age_15_below_percentage, 1),
        'age_above_15_percentage': round(age_above_15_percentage, 1),
        'total_age_percentage': round(total_age_percentage, 1),
        'total_animal_bite_I': total_animal_bite_I,
        'total_animal_bite_II': total_animal_bite_II,
        'total_animal_bite_III': total_animal_bite_III,
        'total_animal_bite_I_percentage': round(total_animal_bite_I_percentage,1),
        'total_animal_bite_II_percentage': round(total_animal_bite_II_percentage,1),
        'total_animal_bite_III_percentage': round(total_animal_bite_III_percentage,1),
        'total_category_percentage': round(total_category_percentage,1),
        'overall_total':overall_total,
        'total_percent':round(total_percent, ),
        'total_tcv_given': total_tcv_given,
        'total_hrig_given': total_hrig_given,
        'total_rig_given': total_erig_given,
        'total_tcv_percentage': round(total_tcv_percentage, 1),
        'total_hrig_percentage': round(total_hrig_percentage, 1),
        'total_rig_percentage': round(total_rig_percentage, 1),
        'total_tcv_percent':round(total_tcv_percent, ),
        'total_erig_percent':round(total_erig_percent, ),
        'total_hrig_percent':round(total_hrig_percent, ),
        'total_dog_bites': total_dog_bites,
        'total_cat_bites': total_cat_bites,
        'total_other_bites': total_other_bites,
        'dog_bite_percentage': round(dog_bite_percentage, 1),
        'cat_bite_percentage': round(cat_bite_percentage, 1),
        'other_bite_percentage': round(other_bite_percentage, 1),
        'total_animal_type_percentage': round(total_animal_type_percentage, 1),
        'total_tcv_given': sum(entry.get('total_tcv_given', 0) for entry in data),
        'total_hrig_given': sum(entry.get('total_hrig_given', 0) for entry in data),
        'total_erig_given': sum(entry.get('total_erig_given', 0) for entry in data),
        'total_immunized': total_immunized,
        'total_unimmunized': total_unimmunized,
        'immunized_percentage': round(immunized_percentage, 1), 
        'unimmunized_percentage': round(unimmunized_percentage, 1),  
        'total_human_rabies':total_human_rabies,
    }

    return render(request, 'monitoring/report.html', context)




@login_required
def tables(request):
    user = request.user

    # Fetch all municipalities
    all_municipalities = Municipality.objects.all()

    # Filter municipalities based on the user
    if user.is_superuser:
        municipalities = Municipality.objects.filter(patients_muni__histories__isnull=False).distinct()
    else:
        municipalities = Municipality.objects.filter(
            patients_muni__user=user,
            patients_muni__histories__isnull=False
        ).distinct()

    # Get the selected municipality and barangay search from the POST parameters
    if request.method == 'POST':
        selected_municipality = request.POST.get('municipality', None)
        barangay_search = request.POST.get('barangay', '').strip()  # Get barangay search input and remove extra spaces
    else:
        selected_municipality = "ALL"
        barangay_search = ""  # Default to an empty search

    # Initialize variables for summary
    total_barangays = 0
    total_patients = 0
    barangay_summary = []
    municipality_summary = []

    if selected_municipality == "ALL" or selected_municipality is None:
        # If "ALL" is selected, show the municipality-level summary
        if user.is_superuser:
            municipality_summary = (
                History.objects.values('muni_id__muni_name')  # Group by municipality
                .annotate(total_barangays=Count('brgy_id', distinct=True),  # Count distinct barangays
                          total_cases=Count('history_id'))  # Count total cases (patients)
                .order_by('muni_id__muni_name')
            )
        else:
            municipality_summary = (
                History.objects.filter(patient_id__user=user)  # Only histories for this user
                .values('muni_id__muni_name')
                .annotate(total_barangays=Count('brgy_id', distinct=True),  # Count distinct barangays
                          total_cases=Count('history_id'))  # Count total cases (patients)
                .order_by('muni_id__muni_name')
            )

        # Calculate the total number of barangays and patients across all municipalities
        total_barangays = sum(record['total_barangays'] for record in municipality_summary)
        total_patients = sum(record['total_cases'] for record in municipality_summary)

    else:
        # Get all barangays for the selected municipality
        barangays_in_municipality = Barangay.objects.filter(muni_id__muni_name=selected_municipality)
        
        # Filter barangays based on the search input (partial match)
        if barangay_search:
            barangays_in_municipality = barangays_in_municipality.filter(brgy_name__icontains=barangay_search)

        # Get barangay-level case summary
        if user.is_superuser:
            cases_in_barangays = History.objects.filter(muni_id__muni_name=selected_municipality) \
                .values('brgy_id__brgy_name') \
                .annotate(total_patients=Count('history_id')) \
                .order_by('brgy_id__brgy_name')
        else:
            cases_in_barangays = History.objects.filter(
                patient_id__user=user,
                muni_id__muni_name=selected_municipality
            ).values('brgy_id__brgy_name') \
                .annotate(total_patients=Count('history_id')) \
                .order_by('brgy_id__brgy_name')

        # Create a dictionary for easy lookup
        cases_dict = {record['brgy_id__brgy_name']: record['total_patients'] for record in cases_in_barangays}

        # Populate barangay summary: if no cases, set total patients to 0
        barangay_summary = [
            {
                'brgy_name': barangay.brgy_name,
                'total_patients': cases_dict.get(barangay.brgy_name, 0)  # Get patients, default to 0 if not found
            }
            for barangay in barangays_in_municipality
        ]

        # Calculate totals
        total_barangays = barangays_in_municipality.count()
        total_patients = sum(record['total_patients'] for record in barangay_summary)

        # Summary for the selected municipality
        municipality_summary = [
            {
                'muni_id__muni_name': selected_municipality,
                'total_barangays': total_barangays,
                'total_cases': total_patients,
            }
        ]

    # Context to pass to the template
    context = {
        'barangay_summary': barangay_summary,
        'municipality_summary': municipality_summary,
        'municipalities': municipalities,
        'selected_municipality': selected_municipality,
        'barangay_search': barangay_search,  # Pass the search input to the template
        'total_barangays': total_barangays,
        'total_patients': total_patients,
        'all_municipalities': all_municipalities,
    }

    return render(request, 'monitoring/table.html', context)





from django.db.models import Q

@login_required
def tables(request):
    user = request.user

    # Fetch all municipalities
    all_municipalities = Municipality.objects.all()

    # Filter municipalities based on the user
    if user.is_superuser:
        municipalities = Municipality.objects.filter(patients_muni__histories__isnull=False).distinct()
    else:
        municipalities = Municipality.objects.filter(
            patients_muni__user=user,
            patients_muni__histories__isnull=False
        ).distinct()

    # Get the selected municipality and barangay search from the POST parameters
    if request.method == 'POST':
        selected_municipality = request.POST.get('municipality', None)
        barangay_search = request.POST.get('barangay', '').strip()  # Get barangay search input and remove extra spaces
    else:
        selected_municipality = "ALL"
        barangay_search = ""  # Default to an empty search

    # Initialize variables for summary
    total_barangays = 0
    total_patients = 0
    barangay_summary = []
    municipality_summary = []

    # If there's a search term for barangay, show barangays matching the search term
    if barangay_search:
        # Get all barangays matching the search term, including abbreviations
        barangays = Barangay.objects.filter(
            Q(brgy_name__icontains=barangay_search) | 
            Q(brgy_name__icontains='Pob') & Q(brgy_name__icontains=barangay_search) |
            Q(brgy_name__icontains='Poblacion') & Q(brgy_name__icontains=barangay_search)
        )

        # Create a summary with the total patients for each barangay across all municipalities
        cases_in_barangays = History.objects.filter(brgy_id__in=barangays) \
            .values('brgy_id__brgy_name', 'muni_id__muni_name') \
            .annotate(total_patients=Count('history_id')) \
            .order_by('muni_id__muni_name', 'brgy_id__brgy_name')

        # Create a dictionary for easy lookup
        cases_dict = {record['brgy_id__brgy_name']: record['total_patients'] for record in cases_in_barangays}

        # Populate barangay summary with corresponding municipalities
        barangay_summary = [
            {
                'muni_name': record['muni_id__muni_name'],
                'brgy_name': record['brgy_id__brgy_name'],
                'total_patients': cases_dict.get(record['brgy_id__brgy_name'], 0)  # Get patients, default to 0 if not found
            }
            for record in cases_in_barangays
        ]

        # Calculate totals
        total_barangays = len(barangay_summary)
        total_patients = sum(record['total_patients'] for record in barangay_summary)

    else:
        # Existing logic for when no search term is provided
        if selected_municipality == "ALL" or selected_municipality is None:
            # If "ALL" is selected, show the municipality-level summary
            if user.is_superuser:
                municipality_summary = (
                    History.objects.values('muni_id__muni_name')  # Group by municipality
                    .annotate(total_barangays=Count('brgy_id', distinct=True),  # Count distinct barangays
                              total_cases=Count('history_id'))  # Count total cases (patients)
                    .order_by('muni_id__muni_name')
                )
            else:
                municipality_summary = (
                    History.objects.filter(patient_id__user=user)  # Only histories for this user
                    .values('muni_id__muni_name')
                    .annotate(total_barangays=Count('brgy_id', distinct=True),  # Count distinct barangays
                              total_cases=Count('history_id'))  # Count total cases (patients)
                    .order_by('muni_id__muni_name')
                )

            # Calculate the total number of barangays and patients across all municipalities
            total_barangays = sum(record['total_barangays'] for record in municipality_summary)
            total_patients = sum(record['total_cases'] for record in municipality_summary)

        else:
            # Get all barangays for the selected municipality
            barangays_in_municipality = Barangay.objects.filter(muni_id__muni_name=selected_municipality)

            # Get barangay-level case summary
            if user.is_superuser:
                cases_in_barangays = History.objects.filter(muni_id__muni_name=selected_municipality) \
                    .values('brgy_id__brgy_name') \
                    .annotate(total_patients=Count('history_id')) \
                    .order_by('brgy_id__brgy_name')
            else:
                cases_in_barangays = History.objects.filter(
                    patient_id__user=user,
                    muni_id__muni_name=selected_municipality
                ).values('brgy_id__brgy_name') \
                    .annotate(total_patients=Count('history_id')) \
                    .order_by('brgy_id__brgy_name')

            # Create a dictionary for easy lookup
            cases_dict = {record['brgy_id__brgy_name']: record['total_patients'] for record in cases_in_barangays}

            # Populate barangay summary: if no cases, set total patients to 0
            barangay_summary = [
                {
                    'brgy_name': barangay.brgy_name,
                    'total_patients': cases_dict.get(barangay.brgy_name, 0)  # Get patients, default to 0 if not found
                }
                for barangay in barangays_in_municipality
            ]

            # Calculate totals
            total_barangays = barangays_in_municipality.count()
            total_patients = sum(record['total_patients'] for record in barangay_summary)

            # Summary for the selected municipality
            municipality_summary = [
                {
                    'muni_id__muni_name': selected_municipality,
                    'total_barangays': total_barangays,
                    'total_cases': total_patients,
                }
            ]

    # Context to pass to the template
    context = {
        'barangay_summary': barangay_summary,
        'municipality_summary': municipality_summary,
        'municipalities': municipalities,
        'selected_municipality': selected_municipality,
        'barangay_search': barangay_search,  # Pass the search input to the template
        'total_barangays': total_barangays,
        'total_patients': total_patients,
        'all_municipalities': all_municipalities,
    }

    return render(request, 'monitoring/table.html', context)








@login_required
def download(request):
    user = request.user

    # Get selected filters from request
    selected_municipality = request.GET.get('municipality')
    selected_barangay = request.GET.get('barangay')
    selected_user = request.GET.get('searchUsername') 
    start_month = request.GET.get('startMonth')
    end_month = request.GET.get('endMonth')
    search_name = request.GET.get('searchName')
    
    # Fetch the histories with related patient, municipality, and barangay data
    histories = History.objects.select_related('patient_id', 'muni_id', 'brgy_id').order_by('-registration_no')

    if not user.is_superuser:
        # Filter histories for the current user if not a superuser
        histories = histories.filter(patient_id__user=user)
    
    patients = Patient.objects.all()

    # Apply filters based on selected municipality and barangay
    if selected_municipality:
        histories = histories.filter(muni_id=selected_municipality)
    if selected_barangay:
        histories = histories.filter(brgy_id=selected_barangay)

    # Apply filter based on username search only if user is a superuser
    if user.is_superuser and selected_user:
        histories = histories.filter(patient_id__user__username=selected_user)
    
    # Fetch unique users from the Patient model for the dropdown if user is superuser
    if user.is_superuser:
        patient_users = Patient.objects.select_related('user').values_list('user', flat=True).distinct()
        users = User.objects.filter(id__in=patient_users)
    else:
        users = []

    # Apply filter based on selected start and end months
    if start_month and end_month:
        try:
            current_year = datetime.now().year
            start_date = datetime.strptime(f"{start_month} {current_year}", "%B %Y").replace(day=1)
            end_date = datetime.strptime(f"{end_month} {current_year}", "%B %Y").replace(day=1) + relativedelta(months=1) - relativedelta(days=1)
            histories = histories.filter(date_registered__gte=start_date, date_registered__lte=end_date)
        except Exception as e:
            print(f"Error parsing date: {e}")  # Log the error and continue

     # Apply filter based on name search
    if search_name:
        histories = histories.filter(Q(patient_id__first_name__icontains=search_name) | Q(patient_id__last_name__icontains=search_name))

    months = month_name[1:]
    
    # Calculate age and attach to each history instance
    for history in histories:
        history.treatment = Treatment.objects.filter(patient_id=history.patient_id).first()
        history.age = calculate_age(history.patient_id.birthday)

    # Count the number of male and female patients
    male = histories.filter(patient_id__sex__iexact='Male').count()
    female = histories.filter(patient_id__sex__iexact='Female').count()

    # Calculate the number of age
    age_15_or_less_count = 0
    age_above_15_count = 0

    for history in histories:
        age = calculate_age(history.patient_id.birthday)
        if age <= 15:
            age_15_or_less_count += 1
        else:
            age_above_15_count += 1

    # Calculate counts for different animal bites   
    source_of_exposure_counter = Counter(histories.values_list('source_of_exposure', flat=True))
    dog_count = source_of_exposure_counter.get('Dog', 0)
    cat_count = source_of_exposure_counter.get('Cat', 0)
    other_animal_count = source_of_exposure_counter.get('Others', 0)

    paginator = Paginator(histories,10)  
    page_number = request.GET.get('page',1)
    try:
        histories = paginator.get_page(page_number)
    except PageNotAnInteger:
        histories = paginator.get_page(1)
    except EmptyPage:
        histories = paginator.get_page(paginator.num_pages)

    # Collect current query parameters
    query_params = request.GET.dict()
    if 'page' in query_params:
        del query_params['page']  # Remove the page parameter
    query_string = urlencode(query_params)

    municipalities = Municipality.objects.all()
    barangays = Barangay.objects.all()


    context = {
        'histories': histories,
        'municipalities': municipalities,
        'barangays': barangays,
        'selected_municipality': selected_municipality,
        'selected_barangay': selected_barangay,
        'selected_user':selected_user,
        'start_month': start_month,
        'end_month': end_month,
        'search_name': search_name,
        'months': months,
        'male' : male,
        'female' : female,
        'dog_count' : dog_count,
        'cat_count' : cat_count,
        'other_animal_count' : other_animal_count,
        'age_15_or_less_count' : age_15_or_less_count,
        'age_above_15_count' : age_above_15_count,
        'query_string' : query_string,
        'users':users
    }

    return render(request, 'monitoring/download.html',context)



@login_required
def download_excel(request):

    
    # Create a workbook and a worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Animal'


    # Save the workbook to a BytesIO stream
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Create the response
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="animal_bite_report.xlsx"'
    return response


def export_excel(request):
    # Create a workbook and get the active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Animal"

    # Define the table headers based on your HTML table structure
    headers = [
        # Row 1
        ["Barangay/ABTC", "Human Case", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "Biting Animals", "", "PEP Coverage", "", "Vaccination Status of Biting Animal", ""],
        # Row 2
        ["", "", "SEX", "", "", "AGE", "", "", "", "Animal Bite", "", "", "HR", "", "Post Exposure Prophylaxis", "", "", "", "", "", "", "", "", "", ""],
        # Row 3
        ["", "", "", "", "", "", "", "", "By Category", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        # Row 4
        ["", "M", "F", "Total", "<=15", ">15", "Total", "I", "II", "III", "Total", "%/Total", "No.", "TCV", "HRIG", "ERIG", "D", "C", "O", "Total", "%TCV", "%ERIG", "Yes", "No"]
    ]

    # Write headers to the sheet
    for row_num, row in enumerate(headers, 1):
        for col_num, value in enumerate(row, 1):
            cell = sheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Align text to center both horizontally and vertically

    # Merge the cells as per colspan and rowspan in the HTML table
    sheet.merge_cells('A1:A4')  # "Barangay/ABTC" merges across 4 rows
    sheet.merge_cells('B1:P1')  # "Human Case" merges from B1 to P1
    sheet.merge_cells('Q1:T1')  # "Biting Animals" merges from Q1 to T1
    sheet.merge_cells('U1:V1')  # "PEP Coverage" merges from U1 to V1
    sheet.merge_cells('W1:X1')  # "Vaccination Status of Biting Animal" merges from W1 to X1

    sheet.merge_cells('B2:D3')  # "SEX" merges from B2 to D3
    sheet.merge_cells('E2:G3')  # "AGE" merges from E2 to G3
    sheet.merge_cells('H2:L2')  # "Animal Bite" merges from H2 to L2
    sheet.merge_cells('M2:M3')  # "HR" merges from M2 to M3
    sheet.merge_cells('N2:P3')  # "Post Exposure Prophylaxis" merges from N2 to P3

    # Individual column merging in the 4th row
    sheet.merge_cells('H3:H4')  # "By Category I"
    sheet.merge_cells('I3:I4')  # "By Category II"
    sheet.merge_cells('J3:J4')  # "By Category III"
    sheet.merge_cells('K3:K4')  # "Total"
    sheet.merge_cells('L3:L4')  # "%/Total"

    sheet.merge_cells('M4:M4')  # "No."
    sheet.merge_cells('N4:N4')  # "TCV"
    sheet.merge_cells('O4:O4')  # "HRIG"
    sheet.merge_cells('P4:P4')  # "ERIG"

    sheet.merge_cells('Q4:Q4')  # "D"
    sheet.merge_cells('R4:R4')  # "C"
    sheet.merge_cells('S4:S4')  # "O"
    sheet.merge_cells('T4:T4')  # "Total"
    sheet.merge_cells('U4:U4')  # "%TCV"
    sheet.merge_cells('V4:V4')  # "%ERIG"
    sheet.merge_cells('W4:W4')  # "Yes"
    sheet.merge_cells('X4:X4')  # "No"

    # Assuming you have data in a similar structure as your HTML
    data = [
        # Example row data
        ['Barangay 1', 10, 5, 15, 7, 8, 15, 3, 6, 1, 10, '100%', 2, 8, 3, 1, 4, 5, 1, 10, '80%', '90%', 8, 2],
        ['Barangay 2', 12, 6, 18, 9, 9, 18, 4, 5, 2, 11, '95%', 1, 7, 2, 1, 3, 6, 1, 10, '85%', '88%', 7, 3],
        # Add more rows as per your data
    ]

    # Write the data to the sheet
    for row_num, row_data in enumerate(data, len(headers) + 1):
        for col_num, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Align all data cells to the center

    # Adjust the column widths for readability
    for i in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(i)].width = 15

    # Set the response to download the file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=animal_bite_report.xlsx'

    # Save the workbook to the response
    workbook.save(response)
    return response

def exp_excel(request):
    # Get the selected quarter from the URL parameters
    selected_quarter = request.GET.get('quarter', '1')  # Default to '1' if no quarter is selected

    # Determine the start and end dates based on the selected quarter
    if selected_quarter == '1':
        start_date = '2024-01-01'
        end_date = '2024-03-31'
    elif selected_quarter == '2':
        start_date = '2024-04-01'
        end_date = '2024-06-30'
    elif selected_quarter == '3':
        start_date = '2024-07-01'
        end_date = '2024-09-30'
    else:
        start_date = '2024-10-01'
        end_date = '2024-12-31'

    # Filter History data by date_registered based on the selected quarter
    data = History.objects.filter(date_registered__range=[start_date, end_date])

    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Q{selected_quarter} Data"

    # Define the headers based on your table structure
    headers = ['Patient Name', 'Date of Exposure', 'Barangay', 'Category', 'Animal Type', 'Bite Site', 'Human Rabies', 'Vaccine Given']
    ws.append(headers)

    # Populate the table rows with filtered data
    for history in data:
        patient = history.patient_id  # Get the patient linked to the history
        treatments = patient.treatments_patient.all()  # Get all treatments for the patient
        
        # Check if there are treatments, and get the first treatment's TCV date if it exists
        vaccine_given = treatments.first().tcv_given if treatments.exists() else 'N/A'

        # Prepare the row with the data
        row = [
            f"{patient.first_name} {patient.last_name}",  # Patient Name
            history.date_of_exposure,  # Date of Exposure
            history.brgy_id.brgy_name,  # Barangay Name
            history.category_of_exposure,  # Category
            history.source_of_exposure,  # Animal Type
            history.bite_site,  # Bite Site
            'Yes' if history.human_rabies else 'No',  # Human Rabies
            vaccine_given  # Vaccine Given (if exists, otherwise N/A)
        ]
        ws.append(row)

    # Set up HTTP response to serve the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=report_Q{selected_quarter}.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response


@login_required
# Function to generate the Excel report
def download_report_excel(request):
    user = request.user
    municipality_map = {
        "MAR": "Maripipi",
        "KAW": "Kawayan",
        "NAV": "Naval",
        "CAIB": "Caibiran"
    }

    selected_quarter = request.GET.get('quarter', '1')
    year = 2024

    quarter_ranges = {
        '1': (date(year, 1, 1), date(year, 3, 31)),
        '2': (date(year, 4, 1), date(year, 6, 30)),
        '3': (date(year, 7, 1), date(year, 9, 30)),
        '4': (date(year, 10, 1), date(year, 12, 31)),
    }

    start_date, end_date = quarter_ranges[selected_quarter]

    # Example data - Replace with your actual data
    data = [
        {
            'barangay': 'Barangay A',
            'data_male': 10,
            'data_female': 12,
            'data_total': 22,
            'age_15_below': 8,
            'age_above_15': 14,
            'total_animal_bite_I': 5,
            'total_animal_bite_II': 4,
            'total_animal_bite_III': 7,
            'total_tcv_given': 5,
            'total_hrig_given': 3,
            'total_erig_given': 2,
            'total_dog_bites': 9,
            'total_cat_bites': 5,
            'total_other_bites': 1,
            'immunized_count': 15,
            'unimmunized_count': 7,
            'human_rabies_count': 0
        },
        # Add more rows as per your actual data
    ]

    # Create an Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Animal Bite and Rabies Report"

    # Set headers
    headers = [
        'Barangay', 'Male', 'Female', 'Total', 'Age 15 and Below', 'Age Above 15',
        'Total Animal Bite I', 'Total Animal Bite II', 'Total Animal Bite III', 
        'Total TCV Given', 'Total HRIG Given', 'Total ERIG Given', 'Dog Bites', 
        'Cat Bites', 'Other Bites', 'Immunized', 'Unimmunized', 'Human Rabies Cases'
    ]
    ws.append(headers)

    # Populate data rows
    for row in data:
        ws.append([ 
            row['barangay'], row['data_male'], row['data_female'], row['data_total'],
            row['age_15_below'], row['age_above_15'], row['total_animal_bite_I'],
            row['total_animal_bite_II'], row['total_animal_bite_III'], 
            row['total_tcv_given'], row['total_hrig_given'], row['total_erig_given'],
            row['total_dog_bites'], row['total_cat_bites'], row['total_other_bites'],
            row['immunized_count'], row['unimmunized_count'], row['human_rabies_count']
        ])

    # Create a BytesIO object to save the Excel file into memory
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)  # Move the pointer to the beginning of the file

    # Create the response object for the Excel download
    response = HttpResponse(
        content=excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="animal_bite_report.xlsx"'

    return response




@login_required
def download_report_pdf(request):
    user = request.user
    municipality_map = {
        "MAR": "Maripipi",
        "KAW": "Kawayan",
        "NAV": "Naval",
        "CAIB": "Caibiran"
    }

    selected_quarter = request.GET.get('quarter', '1')
    year = 2024

    quarter_ranges = {
        '1': (date(year, 1, 1), date(year, 3, 31)),
        '2': (date(year, 4, 1), date(year, 6, 30)),
        '3': (date(year, 7, 1), date(year, 9, 30)),
        '4': (date(year, 10, 1), date(year, 12, 31)),
    }

    start_date, end_date = quarter_ranges[selected_quarter]

    # Example data
    data = [
        {
            'barangay': 'Barangay A',
            'data_male': 10,
            'data_female': 12,
            'age_15_below': 8,
            'age_above_15': 14,
            'total_animal_bite_I': 5,
            'total_animal_bite_II': 4,
            'total_animal_bite_III': 7,
            'total_tcv_given': 5,
            'total_hrig_given': 3,
            'total_erig_given': 2,
            'total_dog_bites': 9,
            'total_cat_bites': 5,
            'total_other_bites': 1,
            'immunized_count': 15,
            'unimmunized_count': 7,
            'human_rabies_count': 0
        },
        # Add more rows as per your actual data
    ]

    # Precompute totals in the view
    total_male = sum(row['data_male'] for row in data)
    total_female = sum(row['data_female'] for row in data)
    total_age_15_below = sum(row['age_15_below'] for row in data)
    total_age_above_15 = sum(row['age_above_15'] for row in data)
    total_animal_bite_I = sum(row['total_animal_bite_I'] for row in data)
    total_animal_bite_II = sum(row['total_animal_bite_II'] for row in data)
    total_animal_bite_III = sum(row['total_animal_bite_III'] for row in data)
    total_tcv_given = sum(row['total_tcv_given'] for row in data)
    total_hrig_given = sum(row['total_hrig_given'] for row in data)
    total_erig_given = sum(row['total_erig_given'] for row in data)
    total_dog_bites = sum(row['total_dog_bites'] for row in data)
    total_cat_bites = sum(row['total_cat_bites'] for row in data)
    total_other_bites = sum(row['total_other_bites'] for row in data)
    total_immunized = sum(row['immunized_count'] for row in data)
    total_unimmunized = sum(row['unimmunized_count'] for row in data)
    total_human_rabies = sum(row['human_rabies_count'] for row in data)

    # Pass totals and data to the template
    html_content = render_to_string('monitoring/report.html', {
        'municipality_name': municipality_map.get(user.code, 'Unknown'),
        'selected_quarter': selected_quarter,
        'data': data,
        'total_male': total_male,
        'total_female': total_female,
        'total_age_15_below': total_age_15_below,
        'total_age_above_15': total_age_above_15,
        'total_animal_bite_I': total_animal_bite_I,
        'total_animal_bite_II': total_animal_bite_II,
        'total_animal_bite_III': total_animal_bite_III,
        'total_tcv_given': total_tcv_given,
        'total_hrig_given': total_hrig_given,
        'total_erig_given': total_erig_given,
        'total_dog_bites': total_dog_bites,
        'total_cat_bites': total_cat_bites,
        'total_other_bites': total_other_bites,
        'total_immunized': total_immunized,
        'total_unimmunized': total_unimmunized,
        'total_human_rabies': total_human_rabies
    })

    # Create a response object for the PDF download
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Report.pdf"'

    # Generate the PDF from HTML
    pisa_status = pisa.CreatePDF(html_content, dest=response)

    # If the PDF creation fails, return an error message
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html_content + '</pre>', content_type='text/html')

    return response


from io import BytesIO
from django.http import HttpResponse
from django.template.loader import render_to_string
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from django.contrib.auth.decorators import login_required
from datetime import datetime
from collections import Counter
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q
from urllib.parse import urlencode

# PDF Generation Function
@login_required
def download_masterlist_pdf(request):
    # Fetch data similar to your existing download view
    user = request.user
    selected_municipality = request.GET.get('municipality')
    selected_barangay = request.GET.get('barangay')
    selected_user = request.GET.get('searchUsername') 
    start_month = request.GET.get('startMonth')
    end_month = request.GET.get('endMonth')
    search_name = request.GET.get('searchName')

    # Query for histories
    histories = History.objects.select_related('patient_id', 'muni_id', 'brgy_id').order_by('-registration_no')
    
    if selected_municipality:
        histories = histories.filter(muni_id=selected_municipality)
    if selected_barangay:
        histories = histories.filter(brgy_id=selected_barangay)
    if selected_user and user.is_superuser:
        histories = histories.filter(patient_id__user__username=selected_user)
    if start_month and end_month:
        start_date = datetime.strptime(f"{start_month} {datetime.now().year}", "%B %Y").replace(day=1)
        end_date = datetime.strptime(f"{end_month} {datetime.now().year}", "%B %Y").replace(day=1) + relativedelta(months=1) - relativedelta(days=1)
        histories = histories.filter(date_registered__gte=start_date, date_registered__lte=end_date)
    if search_name:
        histories = histories.filter(Q(patient_id__first_name__icontains=search_name) | Q(patient_id__last_name__icontains=search_name))

    # Initialize PDF Response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="rabies_masterlist.pdf"'
    
    # Create PDF using ReportLab
    buffer = BytesIO()
    p = canvas.Canvas(buffer)

    # PDF title and headers
    p.setFont("Helvetica", 12)
    p.drawString(50, 800, "BILIRAN PROVINCIAL HOSPITAL")
    p.drawString(50, 780, "ANIMAL BITE TREATMENT CENTER")
    p.drawString(50, 760, "Rabies Masterlist Report")

    # Add column headers to the PDF
    p.drawString(50, 740, "Registration No.")
    p.drawString(150, 740, "Name")
    p.drawString(300, 740, "Barangay")
    p.drawString(400, 740, "Municipality")
    p.drawString(500, 740, "Sex")
    p.drawString(600, 740, "Date of Exposure")
    
    # Draw patient history rows
    y_position = 720
    for history in histories:
        p.drawString(50, y_position, str(history.registration_no))
        p.drawString(150, y_position, f"{history.patient_id.first_name} {history.patient_id.last_name}")
        p.drawString(300, y_position, history.brgy_id.brgy_name)
        p.drawString(400, y_position, history.muni_id.muni_name)
        p.drawString(500, y_position, history.patient_id.sex)
        p.drawString(600, y_position, str(history.date_of_exposure))
        y_position -= 20  # Move to the next row

    p.showPage()
    p.save()

    # Write the PDF data to the response
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)

    return response


# Excel Generation Function
@login_required
def download_masterlist_excel(request):
    # Fetch data similar to your existing download view
    user = request.user
    selected_municipality = request.GET.get('municipality')
    selected_barangay = request.GET.get('barangay')
    selected_user = request.GET.get('searchUsername') 
    start_month = request.GET.get('startMonth')
    end_month = request.GET.get('endMonth')
    search_name = request.GET.get('searchName')

    # Query for histories
    histories = History.objects.select_related('patient_id', 'muni_id', 'brgy_id').order_by('-registration_no')
    
    if selected_municipality:
        histories = histories.filter(muni_id=selected_municipality)
    if selected_barangay:
        histories = histories.filter(brgy_id=selected_barangay)
    if selected_user and user.is_superuser:
        histories = histories.filter(patient_id__user__username=selected_user)
    if start_month and end_month:
        start_date = datetime.strptime(f"{start_month} {datetime.now().year}", "%B %Y").replace(day=1)
        end_date = datetime.strptime(f"{end_month} {datetime.now().year}", "%B %Y").replace(day=1) + relativedelta(months=1) - relativedelta(days=1)
        histories = histories.filter(date_registered__gte=start_date, date_registered__lte=end_date)
    if search_name:
        histories = histories.filter(Q(patient_id__first_name__icontains=search_name) | Q(patient_id__last_name__icontains=search_name))

    # Create Excel file using openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Masterlist Report"

    # Add headers
    headers = ['Registration No.', 'Name', 'Barangay', 'Municipality', 'Sex', 'Date of Exposure']
    ws.append(headers)

    # Populate data rows
    for history in histories:
        ws.append([
            history.registration_no,
            f"{history.patient_id.first_name} {history.patient_id.last_name}",
            history.brgy_id.brgy_name,
            history.muni_id.muni_name,
            history.patient_id.sex,
            history.date_of_exposure,
        ])

    # Create Excel response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="rabies_masterlist.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    return response

